#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore


DEFAULT_BUSINESS_ID = "papi-hair-design-main"


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def fold_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return ascii_only


def norm_header(value: Any) -> str:
    text = fold_text(str(value or "")).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_email(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip().lower()
    if "@" not in text:
        return None
    local, domain = text.split("@", 1)
    if not local or not domain:
        return None
    local = local.split("+", 1)[0]
    return f"{local}@{domain}"


def normalize_phone(value: str | None) -> str | None:
    if not value:
        return None

    raw = str(value).strip()
    if not raw:
        return None

    keep = []
    plus_used = False
    for idx, ch in enumerate(raw):
        if ch.isdigit():
            keep.append(ch)
            continue
        if ch == "+" and idx == 0 and not plus_used:
            keep.append(ch)
            plus_used = True

    cleaned = "".join(keep)
    if not cleaned:
        return None

    if cleaned.startswith("00"):
        cleaned = f"+{cleaned[2:]}"
    elif cleaned.startswith("0") and len(cleaned) == 10:
        # Slovak local style: 09xx...
        cleaned = f"+421{cleaned[1:]}"
    elif cleaned.isdigit() and len(cleaned) == 12 and cleaned.startswith("421"):
        cleaned = f"+{cleaned}"

    return cleaned


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.replace(microsecond=0).isoformat()

    text = str(value).strip()
    if not text:
        return None

    patterns = [
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]
    for pattern in patterns:
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.replace(microsecond=0).isoformat()
        except ValueError:
            continue
    return None


HEADER_ALIASES: dict[str, set[str]] = {
    "name": {"meno", "name", "full_name", "fullname", "zakaznik", "zakaznik_meno"},
    "email": {"email", "e_mail", "mail"},
    "phone": {"telefon", "telefonne_cislo", "telefon_cislo", "phone", "tel"},
    "last_visit": {"posledna_navsteva", "last_visit", "last_booking"},
    "loyalty": {"vernost", "loyalty", "visits", "pocet_navstev"},
    "no_show": {"neprisiel", "no_show", "noshow"},
    "tags": {"stitky", "tags"},
    "note": {"poznamka", "note", "notes"},
}


@dataclass
class ImportRow:
    source_row: int
    full_name: str
    email: str | None
    phone: str | None
    last_visit_at: str | None
    loyalty: int | None
    no_show: int | None
    tags: str | None
    note: str | None

    def identity_key(self) -> str:
        if self.email:
            return f"email:{self.email}"
        if self.phone:
            return f"phone:{self.phone}"
        folded_name = normalize_spaces(fold_text(self.full_name).lower())
        digest = hashlib.sha1(folded_name.encode("utf-8")).hexdigest()[:16]
        return f"name:{digest}"


def pick_column_indexes(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for idx, header in enumerate(headers):
        normalized = norm_header(header)
        if not normalized:
            continue
        for key, aliases in HEADER_ALIASES.items():
            if normalized in aliases and key not in result:
                result[key] = idx
                break
    if "name" not in result:
        raise ValueError("Missing required column for customer name (expected e.g. 'Meno').")
    return result


def value_at(row: list[Any], index: int | None) -> Any:
    if index is None:
        return None
    if index < 0 or index >= len(row):
        return None
    return row[index]


def workbook_rows(xlsx_path: Path) -> list[ImportRow]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    col = pick_column_indexes(headers)

    rows: list[ImportRow] = []
    for row_index in range(2, ws.max_row + 1):
        raw_values = [ws.cell(row=row_index, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]

        raw_name = normalize_spaces(str(value_at(raw_values, col["name"]) or ""))
        if not raw_name:
            continue

        item = ImportRow(
            source_row=row_index,
            full_name=raw_name,
            email=normalize_email(str(value_at(raw_values, col.get("email")) or "")),
            phone=normalize_phone(str(value_at(raw_values, col.get("phone")) or "")),
            last_visit_at=parse_date(value_at(raw_values, col.get("last_visit"))),
            loyalty=parse_int(value_at(raw_values, col.get("loyalty"))),
            no_show=parse_int(value_at(raw_values, col.get("no_show"))),
            tags=normalize_spaces(str(value_at(raw_values, col.get("tags")) or "")) or None,
            note=normalize_spaces(str(value_at(raw_values, col.get("note")) or "")) or None,
        )
        rows.append(item)

    return rows


def dedupe_rows(rows: list[ImportRow]) -> list[ImportRow]:
    merged: dict[str, ImportRow] = {}
    for row in rows:
        key = row.identity_key()
        existing = merged.get(key)
        if not existing:
            merged[key] = row
            continue

        # Keep latest visit and richer metadata.
        if row.last_visit_at and (not existing.last_visit_at or row.last_visit_at > existing.last_visit_at):
            existing.last_visit_at = row.last_visit_at
        if row.loyalty is not None:
            existing.loyalty = max(existing.loyalty or 0, row.loyalty)
        if row.no_show is not None:
            existing.no_show = max(existing.no_show or 0, row.no_show)
        if row.tags and not existing.tags:
            existing.tags = row.tags
        if row.note and not existing.note:
            existing.note = row.note
        if row.phone and not existing.phone:
            existing.phone = row.phone
        if row.email and not existing.email:
            existing.email = row.email

    return list(merged.values())


def init_firestore(project_id: str | None) -> firestore.Client:
    app = firebase_admin.get_app() if firebase_admin._apps else None  # type: ignore[attr-defined]
    if not app:
        key_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
        if key_path:
            cred = credentials.Certificate(key_path)
            app = firebase_admin.initialize_app(cred, {"projectId": project_id} if project_id else None)
        else:
            app = firebase_admin.initialize_app(options={"projectId": project_id} if project_id else None)
    return firestore.client(app=app)


def import_customers(
    db: firestore.Client,
    business_id: str,
    rows: list[ImportRow],
    apply_changes: bool,
) -> dict[str, Any]:
    now = now_iso()
    customers_ref = db.collection("customers")

    existing_snap = customers_ref.where("business_id", "==", business_id).get()
    by_email: dict[str, firestore.DocumentSnapshot] = {}
    by_phone: dict[str, firestore.DocumentSnapshot] = {}
    by_identity: dict[str, firestore.DocumentSnapshot] = {}
    for snap in existing_snap:
        data = snap.to_dict() or {}
        email = normalize_email(data.get("email"))
        phone = normalize_phone(data.get("phone"))
        identity = data.get("bookio_identity_key")
        if email:
            by_email[email] = snap
        if phone:
            by_phone[phone] = snap
        if isinstance(identity, str) and identity:
            by_identity[identity] = snap

    created = 0
    updated = 0
    unchanged = 0
    failed = 0
    sample_changes: list[dict[str, Any]] = []

    for row in rows:
        identity_key = row.identity_key()

        target = None
        if row.email and row.email in by_email:
            target = by_email[row.email]
        elif row.phone and row.phone in by_phone:
            target = by_phone[row.phone]
        elif identity_key in by_identity:
            target = by_identity[identity_key]

        payload: dict[str, Any] = {
            "business_id": business_id,
            "full_name": row.full_name,
            "email": row.email,
            "phone": row.phone,
            "bookio_identity_key": identity_key,
            "bookio_last_visit_at": row.last_visit_at,
            "bookio_loyalty": row.loyalty,
            "bookio_no_show": row.no_show,
            "bookio_tags": row.tags,
            "bookio_note": row.note,
            "import_source": "bookio",
        }

        if target is None:
            create_payload = dict(payload)
            create_payload["updated_at"] = now
            create_payload["created_at"] = now
            created += 1
            sample_changes.append({"action": "create", "row": row.source_row, "identity": identity_key})
            if apply_changes:
                try:
                    doc_ref = customers_ref.document()
                    doc_ref.set(create_payload)
                    snap = doc_ref.get()
                    if row.email:
                        by_email[row.email] = snap
                    if row.phone:
                        by_phone[row.phone] = snap
                    by_identity[identity_key] = snap
                except Exception:
                    failed += 1
                    created -= 1
            continue

        existing_data = target.to_dict() or {}
        comparable_existing = {
            key: existing_data.get(key)
            for key in payload.keys()
        }
        if comparable_existing == payload:
            unchanged += 1
            continue

        updated += 1
        sample_changes.append({"action": "update", "row": row.source_row, "id": target.id, "identity": identity_key})
        if apply_changes:
            try:
                update_payload = dict(payload)
                update_payload["updated_at"] = now
                target.reference.set(update_payload, merge=True)
            except Exception:
                failed += 1
                updated -= 1

    return {
        "business_id": business_id,
        "apply_changes": apply_changes,
        "total_rows": len(rows),
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "failed": failed,
        "sample_changes": sample_changes[:25],
    }


def write_report(report_dir: Path, report: dict[str, Any]) -> Path:
    report_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = report_dir / f"bookio-import-{stamp}.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import customers from Bookio XLSX to Firestore.")
    parser.add_argument("--xlsx", required=True, help="Path to Bookio XLSX export file.")
    parser.add_argument("--business-id", default=DEFAULT_BUSINESS_ID, help="Target business_id in Firestore.")
    parser.add_argument("--project-id", default=None, help="Firebase project ID override.")
    parser.add_argument("--apply", action="store_true", help="Apply changes. Without this flag, runs in dry-run mode.")
    parser.add_argument(
        "--parse-only",
        action="store_true",
        help="Only parse and dedupe XLSX rows; skip Firestore access.",
    )
    parser.add_argument(
        "--report-dir",
        default="docs/import-reports",
        help="Directory for import report JSON.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"ERROR: XLSX file not found: {xlsx_path}")
        return 1

    try:
        parsed_rows = workbook_rows(xlsx_path)
        deduped_rows = dedupe_rows(parsed_rows)
    except Exception as exc:
        print(f"ERROR: Failed to parse workbook: {exc}")
        return 1

    if args.parse_only:
        report = {
            "business_id": args.business_id,
            "apply_changes": False,
            "parse_only": True,
            "total_rows": len(deduped_rows),
            "created": 0,
            "updated": 0,
            "unchanged": 0,
            "failed": 0,
            "sample_changes": [
                {"action": "parsed", "row": row.source_row, "identity": row.identity_key()}
                for row in deduped_rows[:25]
            ],
        }
    else:
        try:
            db = init_firestore(args.project_id)
            report = import_customers(
                db=db,
                business_id=args.business_id,
                rows=deduped_rows,
                apply_changes=bool(args.apply),
            )
        except Exception as exc:
            print(f"ERROR: Firestore import failed: {exc}")
            return 1

    report.update(
        {
            "source_file": str(xlsx_path),
            "source_rows": len(parsed_rows),
            "deduped_rows": len(deduped_rows),
            "generated_at": now_iso(),
        }
    )
    report_path = write_report(Path(args.report_dir), report)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    print(f"Report written to: {report_path}")

    if args.parse_only:
        print("Parse-only run. No Firestore reads/writes were performed.")
    elif not args.apply:
        print("Dry run only. Re-run with --apply to write to Firestore.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
